"""
JPX 後場 取引代金概算 トラッカー
毎日16:10以降に実行し、プライム・スタンダード・グロース市場の取引代金をExcelに記録する
"""

import sys
import io
import re
import requests
import pdfplumber
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime, date
from pathlib import Path

# ---- 設定 ----
EXCEL_PATH = Path(__file__).parent / "jpx_trading_value.xlsx"
BASE_URL = "https://www.jpx.co.jp/markets/equities/volume-and-value/tvdivq000000derc-att"

# PDFのワードリストの中でプライム市場の名前が出る2回目以降が取引代金セクション
PRIME_LABEL    = "内国株式・プライム市場"
STANDARD_LABEL = "内国株式・スタンダード市場"
GROWTH_LABEL   = "内国株式・グロース市場"


def download_pdf(target_date: date) -> bytes:
    """後場PDFをダウンロードして返す"""
    date_str = target_date.strftime("%Y%m%d")
    url = f"{BASE_URL}/2_{date_str}.pdf"
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}
    resp = requests.get(url, headers=headers, timeout=30)
    if resp.status_code == 404:
        # 祝日・市場休場日はPDFが存在しない → スキップ扱い（exit 0）
        print(f"[SKIP] PDFが存在しません（祝日・休場日の可能性）: {url}")
        sys.exit(0)
    if resp.status_code != 200:
        raise RuntimeError(f"PDFダウンロード失敗: HTTP {resp.status_code} / {url}")
    return resp.content


def extract_trading_values(pdf_bytes: bytes) -> dict:
    """
    PDFから取引代金概算（立会内）を抽出する
    戻り値: {"prime": int, "standard": int, "growth": int}  単位: 百万円
    """
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        words = pdf.pages[0].extract_words()

    # ワードリストから (text, y座標) を対応付け
    # プライム市場名が2回目に出現する行のy座標を取引代金セクションとして使う
    prime_ys = [w["top"] for w in words if w["text"] == PRIME_LABEL]
    # 市場名は3回出現（メインテーブル・売買高概算・取引代金概算）
    if len(prime_ys) < 3:
        raise ValueError(f"取引代金概算セクションが見つかりません（出現回数={len(prime_ys)}）")

    values = {}
    for label_text, key in [
        (PRIME_LABEL,    "prime"),
        (STANDARD_LABEL, "standard"),
        (GROWTH_LABEL,   "growth"),
    ]:
        # 3回目の出現位置が取引代金概算セクション
        occurrences = [i for i, w in enumerate(words) if w["text"] == label_text]
        if len(occurrences) < 3:
            raise ValueError(f"ラベルが3回見つかりません: {label_text}")
        idx = occurrences[2]
        # そのラベルの直後にある数値ワードを取得
        y = words[idx]["top"]
        num_word = next(
            (w for w in words[idx+1:] if abs(w["top"] - y) < 3 and re.match(r"[\d,]+$", w["text"])),
            None,
        )
        if num_word is None:
            raise ValueError(f"数値が見つかりません: {label_text}")
        values[key] = int(num_word["text"].replace(",", ""))

    return values


def load_or_create_workbook():
    """Excelファイルを開くか新規作成する"""
    if EXCEL_PATH.exists():
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "取引代金概算"
        # ヘッダー行
        headers = ["日付", "プライム市場\n（百万円）", "スタンダード市場\n（百万円）", "グロース市場\n（百万円）"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E79")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 36
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 22
        ws.column_dimensions["D"].width = 18
    return wb, ws


def find_last_row(ws):
    """データが入っている最終行を返す（ヘッダー除く）"""
    last = ws.max_row
    while last > 1 and ws.cell(row=last, column=1).value is None:
        last -= 1
    return last


def write_to_excel(target_date: date, values: dict):
    """
    Excelに1行書き込む。
    月が変わった場合は空行を1行挿入してから書く。
    同日付のデータが既にある場合は上書きする。
    """
    wb, ws = load_or_create_workbook()
    last_row = find_last_row(ws)

    # 同日付が既に存在するか確認
    for row in range(2, last_row + 1):
        cell_val = ws.cell(row=row, column=1).value
        if isinstance(cell_val, (date, datetime)):
            existing_date = cell_val.date() if isinstance(cell_val, datetime) else cell_val
            if existing_date == target_date:
                # 上書き
                _write_row(ws, row, target_date, values)
                wb.save(EXCEL_PATH)
                print(f"[UPDATE] {target_date} のデータを上書きしました")
                return

    # 新規行
    new_row = last_row + 1

    # 月替わりチェック: 前行（空行以外）と月が異なれば空行挿入
    if last_row >= 2:
        prev_val = ws.cell(row=last_row, column=1).value
        if prev_val is not None:
            prev_date = prev_val.date() if isinstance(prev_val, datetime) else prev_val
            if prev_date.month != target_date.month:
                new_row = last_row + 2  # 空行を1行分空ける

    _write_row(ws, new_row, target_date, values)

    # 交互背景色
    fill_color = "D6E4F0" if (new_row % 2 == 0) else "FFFFFF"
    for col in range(1, 5):
        ws.cell(row=new_row, column=col).fill = PatternFill("solid", fgColor=fill_color)

    wb.save(EXCEL_PATH)
    print(f"[OK] {target_date} のデータを記録: プライム={values['prime']:,} / スタンダード={values['standard']:,} / グロース={values['growth']:,} （百万円）")


THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _write_row(ws, row: int, target_date: date, values: dict):
    """指定行にデータを書き込む"""
    cells = [
        ws.cell(row=row, column=1, value=target_date),
        ws.cell(row=row, column=2, value=values["prime"]),
        ws.cell(row=row, column=3, value=values["standard"]),
        ws.cell(row=row, column=4, value=values["growth"]),
    ]
    cells[0].number_format = "YYYY/MM/DD"
    cells[0].alignment = Alignment(horizontal="center")
    for c in cells[1:]:
        c.number_format = "#,##0"
        c.alignment = Alignment(horizontal="right")
    for c in cells:
        c.border = THIN_BORDER


def main(target_date: date = None):
    if target_date is None:
        target_date = date.today()

    print(f"[INFO] {target_date} の後場データを取得中...")
    try:
        pdf_bytes = download_pdf(target_date)
        values = extract_trading_values(pdf_bytes)
        write_to_excel(target_date, values)
    except Exception as e:
        print(f"[ERROR] {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    # コマンドライン引数で日付を指定可能: python jpx_scraper.py 20260312
    if len(sys.argv) > 1:
        d = datetime.strptime(sys.argv[1], "%Y%m%d").date()
    else:
        d = date.today()
    main(d)
